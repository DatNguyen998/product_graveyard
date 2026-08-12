#!/usr/bin/env python3
"""
build_graveyard.py — Gộp các "product graveyard" của Big Tech thành 1 dataset chuẩn hoá.

Nguồn (tất cả open-source, MIT/公开):
  Google     codyogden/killedbygoogle        graveyard.json
  Microsoft  fabianoriccardi/killed-by-microsoft  graveyard.json
  Apple      TheDen/killed-by-apple          products.json

Chạy:
  python3 build_graveyard.py                 # xuất JSON + CSV
  python3 build_graveyard.py --xlsx          # thêm Excel (cần: pip install openpyxl)

Output:
  graveyard_all.json / graveyard_all.csv / graveyard_all.xlsx
"""

import argparse
import csv
import json
import sys
import urllib.request
from datetime import date, datetime

# --------------------------------------------------------------------------
# Cấu hình nguồn. Dùng cdn.jsdelivr.net thay vì raw.githubusercontent vì
# jsdelivr có CDN cache + gửi CORS header (dùng lại được cho bản HTML).
# --------------------------------------------------------------------------
SOURCES = [
    {
        "company": "Google",
        "url": "https://cdn.jsdelivr.net/gh/codyogden/killedbygoogle@main/graveyard.json",
        "parser": "kbg",
        "credit": "killedbygoogle.com",
    },
    {
        "company": "Microsoft",
        "url": "https://cdn.jsdelivr.net/gh/fabianoriccardi/killed-by-microsoft@main/graveyard.json",
        "parser": "kbg",  # cùng schema với Killed by Google
        "credit": "killedbymicrosoft.info",
    },
    {
        "company": "Apple",
        "url": "https://cdn.jsdelivr.net/gh/TheDen/killed-by-apple@main/products.json",
        "parser": "kba",
        "credit": "killedbyapple.theden.sh",
    },
]

TODAY = date.today()

# Schema đích — mọi nguồn phải map về đây
FIELDS = [
    "company",
    "name",
    "type",
    "date_open",
    "date_close",
    "year_open",
    "year_close",
    "date_precision",   # "day" | "year"  -> quan trọng: Apple chỉ có năm
    "lifespan_years",
    "status",           # "dead" | "scheduled"  -> có mục dateClose ở tương lai
    "description",
    "link",
    "source",
]


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def fetch_json(url):
    req = urllib.request.Request(
        url, headers={"User-Agent": "graveyard-aggregator/1.0", "Accept": "application/json"}
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode("utf-8"))


def parse_date(s):
    """'2019-04-02' -> date. Trả None nếu không parse được."""
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except ValueError:
            continue
    return None


def years_between(d1, d2):
    if not d1 or not d2:
        return None
    return round((d2 - d1).days / 365.25, 1)


def clean(s):
    return " ".join(str(s or "").split())


# --------------------------------------------------------------------------
# Parsers — mỗi nguồn 1 hàm, trả về list dict theo FIELDS
# --------------------------------------------------------------------------
def parse_kbg(raw, company, credit):
    """Schema Killed by Google / Killed by Microsoft."""
    out = []
    for it in raw:
        d_open = parse_date(it.get("dateOpen"))
        d_close = parse_date(it.get("dateClose"))
        out.append(
            {
                "company": company,
                "name": clean(it.get("name")),
                "type": clean(it.get("type")).lower() or None,
                "date_open": d_open.isoformat() if d_open else None,
                "date_close": d_close.isoformat() if d_close else None,
                "year_open": d_open.year if d_open else None,
                "year_close": d_close.year if d_close else None,
                "date_precision": "day",
                "lifespan_years": years_between(d_open, d_close),
                "status": "scheduled" if (d_close and d_close > TODAY) else "dead",
                "description": clean(it.get("description")),
                "link": clean(it.get("link")),
                "source": credit,
            }
        )
    return out


def parse_kba(raw, company, credit):
    """Schema Killed by Apple — chỉ có năm (born/died), không có ngày."""
    out = []
    items = raw.get("products", raw) if isinstance(raw, dict) else raw
    for it in items:
        y_open, y_close = it.get("born"), it.get("died")
        d_open = parse_date(f"{y_open}-01-01") if y_open else None
        d_close = parse_date(f"{y_close}-12-31") if y_close else None
        refs = it.get("refs") or []
        out.append(
            {
                "company": company,
                "name": clean(it.get("name")),
                "type": ", ".join(it.get("cats") or []).lower() or None,
                "date_open": None,   # cố ý để trống: nguồn không có ngày
                "date_close": None,
                "year_open": y_open,
                "year_close": y_close,
                "date_precision": "year",
                "lifespan_years": (y_close - y_open) if (y_open and y_close) else None,
                "status": "scheduled" if (y_close and y_close > TODAY.year) else "dead",
                "description": clean(it.get("desc")),
                "link": clean(refs[0].get("url")) if refs else "",
                "source": credit,
            }
        )
    return out


PARSERS = {"kbg": parse_kbg, "kba": parse_kba}


# --------------------------------------------------------------------------
# Pipeline
# --------------------------------------------------------------------------
def build():
    rows, report = [], []
    for src in SOURCES:
        try:
            raw = fetch_json(src["url"])
            parsed = PARSERS[src["parser"]](raw, src["company"], src["credit"])
            rows.extend(parsed)
            report.append((src["company"], len(parsed), "ok"))
        except Exception as e:  # noqa: BLE001 — muốn 1 nguồn chết không kéo cả pipeline
            report.append((src["company"], 0, f"FAIL: {e}"))

    # Dedupe theo (company, name) — giữ bản có nhiều thông tin nhất
    seen = {}
    for r in rows:
        key = (r["company"], r["name"].lower())
        if key not in seen or sum(v is not None for v in r.values()) > sum(
            v is not None for v in seen[key].values()
        ):
            seen[key] = r
    deduped = list(seen.values())
    dupes = len(rows) - len(deduped)

    # Sort: mới chết trước
    deduped.sort(key=lambda r: (r["year_close"] or 0, r["date_close"] or ""), reverse=True)
    return deduped, report, dupes


def qa(rows):
    """Kiểm tra chất lượng — in cảnh báo, không tự sửa."""
    problems = []
    for r in rows:
        yo, yc = r["year_open"], r["year_close"]
        if yo and yc and yc < yo:
            problems.append(f"  ! chết trước khi sinh: {r['company']} / {r['name']} ({yo}->{yc})")
        if yo and yo < 1975:
            problems.append(f"  ! năm ra mắt đáng ngờ: {r['company']} / {r['name']} ({yo})")
        if not r["name"]:
            problems.append("  ! thiếu name")
        if not yc:
            problems.append(f"  ! thiếu năm khai tử: {r['company']} / {r['name']}")
    return problems


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", action="store_true", help="xuất thêm file Excel")
    ap.add_argument("--out", default="graveyard_all", help="tên file output (không đuôi)")
    args = ap.parse_args()

    rows, report, dupes = build()

    print("\n=== NGUỒN ===")
    for company, n, status in report:
        print(f"  {company:<12} {n:>4} mục   {status}")
    print(f"  {'TỔNG':<12} {len(rows):>4} mục sau dedupe (bỏ {dupes} trùng)")

    print("\n=== PHÂN BỔ THEO THẬP KỶ (năm khai tử) ===")
    buckets = {}
    for r in rows:
        if r["year_close"]:
            buckets[r["year_close"] // 10 * 10] = buckets.get(r["year_close"] // 10 * 10, 0) + 1
    for dec in sorted(buckets):
        print(f"  {dec}s  {'#' * buckets[dec]} ({buckets[dec]})")

    problems = qa(rows)
    print(f"\n=== QA: {len(problems)} cảnh báo ===")
    for p in problems[:20]:
        print(p)
    if len(problems) > 20:
        print(f"  ... và {len(problems) - 20} cảnh báo nữa")

    with open(f"{args.out}.json", "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    with open(f"{args.out}.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        w.writerows(rows)
    print(f"\n✓ {args.out}.json\n✓ {args.out}.csv")

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "graveyard"
            ws.append(FIELDS)
            for c in ws[1]:
                c.font = Font(bold=True)
            for r in rows:
                ws.append([r[k] for k in FIELDS])
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for i, k in enumerate(FIELDS, 1):
                ws.column_dimensions[get_column_letter(i)].width = 45 if k == "description" else 16
            wb.save(f"{args.out}.xlsx")
            print(f"✓ {args.out}.xlsx")
        except ImportError:
            print("✗ xlsx bỏ qua — chạy: pip install openpyxl", file=sys.stderr)


if __name__ == "__main__":
    main()
